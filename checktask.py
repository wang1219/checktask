# encoding=utf-8

import os
import os.path
import sys
import hashlib
import smtplib
import urllib
import urllib2
import json
import time
import base64
import logging
from logging import handlers
from binascii import b2a_hex
from email.header import Header
from email.mime.text import MIMEText

import xlrd
import requests
from Crypto import Random
from openpyxl import load_workbook
from requests.packages.urllib3.exceptions import InsecureRequestWarning

reload(sys)
sys.setdefaultencoding('utf-8')

# 禁用安全请求警告
requests.packages.urllib3.disable_warnings(InsecureRequestWarning)

# Excel表格所在位置
Excel_filepath = './excel.xls'

# 邮箱配置
MAIL_HOST = 'smtp.sina.com'
MAIL_PORT = 25
MAIL_USER = 'jackchen815@sina.com'
MAIL_PASSWORD = 'qiong***'
MAIL_TO = ['**@126.com']
MAIL_TITLE = u'报表'

# 日志目录
LOG_DIR = '/var/log/checktask'
LOG_LEVEL = 'DEBUG'
LOG_NAME = 'checktask'

# 超时时间设置
CHECK_TIMEOUT = 30 * 60

# DaMaTu相关配置
DamatuUserName = 'xiaolongchang'
DamatuUserPassWD = '858993460'

# 验证码保存路径
PictureFileSavePath = ''


LOG = None

if not os.path.exists(LOG_DIR):
    os.makedirs(LOG_DIR)

if LOG_NAME not in logging.Logger.manager.loggerDict:
    log_file_path = os.path.join(LOG_DIR, '%s.log' % LOG_NAME)
    fh = handlers.TimedRotatingFileHandler(log_file_path, when='d')
    formatter = logging.Formatter(fmt='%(asctime)s %(levelname)-8s '
                                  '%(module)s %(lineno)d %(message)s')
    fh.setFormatter(formatter)
    fh.setLevel(LOG_LEVEL.upper())

    logger = logging.getLogger(LOG_NAME)
    logger.addHandler(fh)
    logger.setLevel(LOG_LEVEL.upper())
    LOG = logger
else:
    LOG = logging.getLogger(LOG_NAME)

HTML_TEMPLATE = """
<html>
<head>
<meta http-equiv = "Content-Type" content = "text/html; charset = UTF-8">
<title>报表</title>
<style type="text/css">
    table {
            border: 1px solid #345;
            border-collapse: collapse;
    }
    table th {
        color: white;
        background-color: #a2c4c9;
    }
    table td, table th {
        padding: 0.5em 1em;
        border: 1px solid #345;
        vertical-align:top;
    }
    table td .index{
        width: 50px;
    }
    table td .hostname{
        width: 200px;
    }
    .summary-table th {
        text-align: right;
    }
    .warning {
        font-weight: bold;
        color: red;
    }
    table td h2 {
        color:blue;
        font-size:1.5em;
        text-align:center;
    }
    .manager {
        color:red;
    }
    .flag {
        font-weight: bold;
        font-size:0.7em;
        color:blue;
    }
</style>
</head>
<body>
    <p>
    <h3>检测结果详情：</h3>
    </p>
    {{report}}
    </br>
</body>
</html>
"""


def md5str(str):  # md5加密字符串
    m = hashlib.md5(str.encode(encoding="utf-8"))
    return m.hexdigest()


def md5(byte):  # md5加密byte
    return hashlib.md5(byte).hexdigest()


class DamatuApi():
    ID = '40838'
    KEY = 'ca9507e17e8d5ddf7c57cd18d8d33010'
    HOST = 'http://api.dama2.com:7766/app/'

    def __init__(self, username, password):
        self.username = username
        self.password = password

    def getSign(self, param=b''):
        return (md5(bytes(self.KEY).encode('utf-8') + bytes(self.username).encode('utf-8') + param))[:8]

    def getPwd(self):
        return md5str(self.KEY + md5str(md5str(self.username) + md5str(self.password)))

    def post(self, path, params={}):
        data = urllib.urlencode(params).encode('utf-8')
        url = self.HOST + path
        response = urllib2.Request(url, data)
        return urllib2.urlopen(response).read()

    # 查询余额 return 是正数为余额 如果为负数 则为错误码
    def getBalance(self):
        data = {'appID': self.ID,
                'user': self.username,
                'pwd': self.getPwd(),
                'sign': self.getSign()
                }
        res = self.post('d2Balance', data)
        res = str(res).encode('utf-8')
        jres = json.loads(res)
        if jres['ret'] == 0:
            return jres['balance']
        else:
            return jres['ret']

    # 上传验证码 参数filePath 验证码图片路径 如d:/1.jpg type是类型，查看http://wiki.dama2.com/index.php?n=ApiDoc.Pricedesc  return 是答案为成功 如果为负数 则为错误码
    def decode(self, filedata, type):
        # f = open(filePath, 'rb')
        # fdata = f.read()
        # filedata = base64.b64encode(fdata)
        # f.close()
        data = {'appID': self.ID,
                'user': self.username,
                'pwd': self.getPwd(),
                'type': type,
                'fileDataBase64': filedata,
                'sign': self.getSign(base64.b64decode(filedata))
                }  # self.getSign(base64.b64decode(filedata))
        res = self.post('d2File', data)
        # res = str(res).encode('utf-8')
        jres = json.loads(res)
        LOG.info('jres', jres)
        if jres['ret'] == 0:
            # 注意这个json里面有ret，id，result，cookie，根据自己的需要获取
            return (jres['result'])
        else:
            return jres['ret']

    # url地址打码 参数 url地址  type是类型(类型查看http://wiki.dama2.com/index.php?n=ApiDoc.Pricedesc) return 是答案为成功 如果为负数 则为错误码
    def decodeUrl(self, url, type):
        data = {'appID': self.ID,
                'user': self.username,
                'pwd': self.getPwd(),
                'type': type,
                'url': urllib.quote(url),
                'sign': self.getSign(url.encode(encoding="utf-8"))
                }
        res = self.post('d2Url', data)
        res = str(res).encode('utf-8')
        jres = json.loads(res)
        if jres['ret'] == 0:
            # 注意这个json里面有ret，id，result，cookie，根据自己的需要获取
            return (jres['result'])
        else:
            return jres['ret']

    # 报错 参数id(string类型)由上传打码函数的结果获得 return 0为成功 其他见错误码
    def reportError(self, id):
        # f=open('0349.bmp','rb')
        # fdata=f.read()
        # print(md5(fdata))
        data = {'appID': self.ID,
                'user': self.username,
                'pwd': self.getPwd(),
                'id': id,
                'sign': self.getSign(id.encode(encoding="utf-8"))
                }
        res = self.post('d2ReportError', data)
        res = str(res).encode('utf-8')
        jres = json.loads(res)
        return jres['ret']


class Bill(object):
    def __init__(self, title, id, num, date, money):
        self.title = title
        self.bill_id = id
        self.bill_num = num
        self.bill_date = date
        self.bill_money = money
        self.verification_code = None
        self.is_ok = False

    def save(self, text):
        with open(PictureFileSavePath, 'w') as f:
            f.write(base64.b64decode(text))
            LOG.info(u'存储验证码')

    def _get_verification_picture(self):
        try:
            LOG.info('查询验证码')
            # callback = "jQuery%s_%s" % (
            #     ''.join([random.choice(string.digits) for _ in range(22)]),
            #     ''.join([random.choice(string.digits) for _ in range(13)]))
            callback = 'jQuery110209791718499773361_1501601178568'
            r = '0.030927983281551885',  # '0.%s' % ''.join([random.choice(string.digits) for _ in range(16)])
            fpdm = self.bill_id
            resp = requests.get('https://fpcyweb.tax.sh.gov.cn:1001/WebQuery/yzmQuery?callback=%s'
                                '&fpdm=%s&r=%s' % (callback, fpdm, r), verify=False)
        except Exception as e:
            LOG.error(u'获取验证码图片错误 - %s' % str(e))
        else:
            return json.loads(resp.content[len(callback) + 1:-1]), callback
        return None, ''

    def check_task(self):
        keys, callback = self._get_verification_picture()
        if keys and keys.get('key4') == '00':
            # keys['key4'] == '00' 为最正常验证码，其他奇葩的他们解析不了
            # self.save(keys['key1'])

            verification_code = dmt.decode(keys['key1'], 200)
            try:
                verification_code = int(verification_code)
                return False
            except ValueError:
                if verification_code == 'ERROR':
                    return False
                self.verification_code = verification_code
                self.is_ok = self._post_request(callback, keys.get('key2'), keys.get('key3'), keys.get('key4'))
                return True

        return False

    def _post_request(self, callback, key2, key3, key4):
        key4 = '01'
        try:
            url = 'https://fpcyweb.tax.sh.gov.cn:1001/WebQuery/query?'
            params = 'callback={callback}&fpdm={fpdm}' \
                     '&fphm={fphm}&kprq={kprq}&fpje={fpje}&fplx={fplx}&yzm={yzm}&yzmSj={yzmSj}&index={index}' \
                     '&iv={iv}&salt={salt}&_={_}'.format(
                callback=callback,
                fpdm=str(self.bill_id),
                fphm=str(self.bill_num),
                kprq=str(self.bill_date),
                fpje=str(self.bill_money),
                fplx=key4,
                yzm=urllib.quote(self.verification_code),
                yzmSj=urllib.quote(key2),
                index=key3,
                iv=b2a_hex(Random.get_random_bytes(128 / 8)),
                salt=b2a_hex(Random.get_random_bytes(128 / 8)),
                _='1501237442168'  # ''.join([random.choice(string.digits) for _ in range(13)])
            )
            url = url + params
            resp = requests.get(url, verify=False)
            status_code = resp.status_code
            LOG.info('status_code:%s' % status_code)
        except Exception as e:
            LOG.error(u'请求验证失败，正在重试请稍等！- %s' % str(e))
            return False
        else:
            if '"key1":"002"' in resp.content:
                LOG.info('URL:%s' % url)
                LOG.info('Find "key1":"002" - %s' % resp.content)
                LOG.info(u'请求达到最大次数')
                return True
        return False

    def __str__(self):
        return "title:%s bill_id:%s %s, bill_num:%s %s, bill_date:%s %s, bill_money:%s %s " % (
            self.title, self.bill_id, type(self.bill_id),
            self.bill_num, type(self.bill_num), self.bill_date,
            type(self.bill_date), self.bill_money, type(self.bill_money)
        )


class Excel(object):
    def setup(self, filepath):
        if not os.path.isfile(filepath):
            raise Exception(u'文件%s不存在' % filepath)

        if filepath.endswith('.xls'):
            self.bills = self._get_excel_xls(filepath)
        else:
            self.bills = self._get_excel_xlsx(filepath)

    def _get_excel_xls(self, filepath):
        wb = xlrd.open_workbook(filepath)
        bills = []
        for sheet in wb.sheets():
            current_title = ''

            for row in range(sheet.nrows):
                row_data = [data for data in sheet.row_values(row)]
                if len(row_data) != 4 or not row_data[0]:
                    continue
                if isinstance(row_data[0], basestring) and \
                        (isinstance(row_data[1], basestring) or
                             isinstance(row_data[2], basestring) or
                             isinstance(row_data[4], basestring)):
                    continue
                if isinstance(row_data[0], basestring):
                    current_title = row_data[0]
                else:
                    bills.append(Bill(current_title, str(row_data[0]), str(row_data[1]),
                                      xlrd.xldate.xldate_as_datetime(row_data[2], 0).strftime("%Y%m%d"), row_data[3]))
        return bills

    def _get_excel_xlsx(self, filepath):
        wb = load_workbook(filepath)
        bills = []
        for sheetname in wb.sheetnames:
            sheet = wb.get_sheet_by_name(sheetname)

            # 按行读取
            current_title = ''
            for i in range(1, sheet.max_row + 1):
                row_data = [col.value for col in sheet[str(i)]]
                if len(row_data) != 4 or not row_data[0]:
                    continue
                if isinstance(row_data[0], basestring) and \
                        (isinstance(row_data[1], basestring) or
                             isinstance(row_data[2], basestring) or
                             isinstance(row_data[4], basestring)):
                    continue
                if isinstance(row_data[0], basestring):
                    current_title = row_data[0]
                else:
                    bills.append(Bill(current_title, int(row_data[0]), int(row_data[1]),
                                      row_data[2].strftime("%Y%m%d"), row_data[3]))
        return bills

    def check_succ(self):
        for bill in self.bills:
            if not bill.is_ok:
                return False
        return True

    def _build_mail_table(self, bills):
        report_list = []
        report_table_head = '<table class="detail-table"><tr><th>序号</th><th>发票代码</th><th>发票号码</th><th>发票日期</th><th>开具金额</th><th>达到最大次数</th></tr>{{report_data}}</table>'

        num = 0
        for bill in bills:
            num += 1
            report_list.append(
                '<tr><td class="index">%s</td><td class="hostname">%s</td><td class="hostname">%s</td><td class="hostname">%s</td><td class="hostname">%s</td><td class="hostname">%s</td></tr>' % (
                    str(num),
                    bill.bill_id,
                    bill.bill_num,
                    bill.bill_date,
                    bill.bill_money,
                    bill.is_ok
                ))
        report_table_head = report_table_head.decode('utf-8')
        return report_table_head.replace('{{report_data}}', ''.join(report_list))

    def _get_report(self):
        org_bill = {}
        report_list = []
        for bill in self.bills:
            if bill.title not in org_bill:
                org_bill[bill.title] = []
            org_bill[bill.title].append(bill)

        for title, bills in org_bill.items():
            report_list.append('<h4>%s：</h4></br>%s</br></br>'.decode('utf-8') %
                               (title.decode('utf-8'), self._build_mail_table(bills)))

        content = HTML_TEMPLATE.decode('utf-8')
        content = content.replace('{{report}}', ''.join(report_list))
        return content

    def _send_mail(self):
        content = self._get_report()
        msg = MIMEText(content, 'html', 'utf-8')
        msg['Subject'] = Header('{sub}'.format(
            sub=MAIL_TITLE.encode('utf-8') if MAIL_TITLE else ''), 'utf-8')
        msg['From'] = MAIL_USER
        msg['To'] = ','.join(MAIL_TO)
        try:
            s = smtplib.SMTP(MAIL_HOST, MAIL_PORT)
            s.login(MAIL_USER, MAIL_PASSWORD)
            s.sendmail(MAIL_USER, MAIL_TO, msg.as_string())
            s.close()
            LOG.info('Send mail succ - %s' % (MAIL_TO))
            return True
        except Exception as e:
            LOG.error('Send mail error - %s - %s - %s' % (
                str(e), MAIL_TITLE, MAIL_TO
            ))
            return False

    def check(self):
        if not hasattr(self, 'bills'):
            LOG.error(u'请先执行setup...')
            return

        num = 0
        LOG.info('Total:%s' % len(self.bills))
        for bill in self.bills:
            is_ok = False
            num += 1
            LOG.info('.............Num: %s.........' % num)
            start = time.time()
            while not is_ok and int(time.time() - start) < CHECK_TIMEOUT:
                try:
                    bill.check_task()
                    is_ok = bill.is_ok
                    LOG.info('is_ok:%s' % is_ok)
                except Exception as e:
                    LOG.error('Check error - %s' % str(e))
                    continue

        self._send_mail()


dmt = DamatuApi(DamatuUserName, DamatuUserPassWD)

if __name__ == '__main__':
    excel = Excel()
    excel.setup(Excel_filepath)
    excel.check()
